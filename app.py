import json
import os
import openpyxl
import requests
from flask import Flask, jsonify, render_template, request

app = Flask(__name__)

SERVICES_KEY = os.environ.get("VIETMAP_SERVICES_KEY", "e3464a9335a846e985861bdf43fd8700201a93af28006a40")
TILEMAP_KEY  = os.environ.get("VIETMAP_TILEMAP_KEY",  "06fadcaa43886a1b8a3fd81709a1f9723bb3e25d1010554b")
EXCEL_PATH   = os.path.join(os.path.dirname(__file__), "Book1.xlsx")

# Average car speed in HCM city (km/h) — used for travel time estimate
HCM_AVG_SPEED_KMH = 25


def load_stores():
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active
    stores = []
    for r in range(2, ws.max_row + 1):
        pos_code        = ws.cell(r, 1).value
        pos_name        = ws.cell(r, 2).value
        full_address    = ws.cell(r, 3).value
        total_order_30d = ws.cell(r, 4).value
        total_order_90d = ws.cell(r, 5).value
        location        = ws.cell(r, 6).value

        if not location or not pos_name:
            continue

        try:
            lat_str, lng_str = str(location).split(",")
            lat = float(lat_str.strip())
            lng = float(lng_str.strip())
        except Exception:
            continue

        stores.append({
            "pos_code":        pos_code or "",
            "pos_name":        pos_name or "",
            "full_address":    full_address or "",
            "total_order_30d": int(total_order_30d) if total_order_30d else 0,
            "total_order_90d": int(total_order_90d) if total_order_90d else 0,
            "lat":             lat,
            "lng":             lng,
        })
    return stores


@app.route("/")
def index():
    return render_template("map.html", tilemap_key=TILEMAP_KEY)


@app.route("/api/stores")
def api_stores():
    stores = load_stores()
    return jsonify({"stores": stores})


@app.route("/api/route")
def api_route():
    """
    Proxy VietMap routing API.
    Params: lat1, lng1, lat2, lng2
    Returns: distance_km, duration_min (based on HCM avg speed), polyline points
    """
    try:
        lat1 = float(request.args["lat1"])
        lng1 = float(request.args["lng1"])
        lat2 = float(request.args["lat2"])
        lng2 = float(request.args["lng2"])
    except (KeyError, ValueError):
        return jsonify({"error": "Missing or invalid lat1/lng1/lat2/lng2"}), 400

    # VietMap routing API (GraphHopper-based)
    params = [
        ("apikey", SERVICES_KEY),
        ("point",  f"{lat1},{lng1}"),
        ("point",  f"{lat2},{lng2}"),
        ("vehicle", "car"),
        ("points_encoded", "true"),
        ("locale", "vi"),
    ]

    try:
        r = requests.get(
            "https://maps.vietmap.vn/api/route",
            params=params,
            timeout=15,
        )
        r.raise_for_status()
        data = r.json()
    except Exception as e:
        return jsonify({"error": str(e)}), 502

    if "paths" not in data or not data["paths"]:
        return jsonify({"error": "No route found"}), 404

    path = data["paths"][0]
    distance_m  = path.get("distance", 0)       # metres
    distance_km = round(distance_m / 1000, 1)

    # Use HCM average speed for travel time
    duration_min = round(distance_km / HCM_AVG_SPEED_KMH * 60)

    return jsonify({
        "distance_km":  distance_km,
        "duration_min": duration_min,
        "points":       path.get("points", ""),  # encoded polyline
    })


if __name__ == "__main__":
    print("Starting store map server at http://localhost:5000")
    app.run(debug=True, port=5000)
